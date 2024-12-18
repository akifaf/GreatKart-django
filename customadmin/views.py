from decimal import Decimal
import json
import os
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import  get_object_or_404, render, redirect
from django.contrib import messages, auth
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.db.models import Sum, Count
from django.db.models.functions import ExtractWeek, ExtractMonth, ExtractYear, ExtractDay

from accounts.models import Account
from orders import models
from orders.models import Order, OrderProduct, Refund
from store.forms import ImageForm
from store.models import Color, Coupon, Product, ProductGallery, Size, Variation
from category.models import Category
from .forms import CategoryForm, OrderForm, ProductForm, VariationForm
from customadmin import forms
from django.contrib.auth.decorators import user_passes_test

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def user_is_admin(user):
    return  user.is_staff  
 
@never_cache
def admin_login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        user_obj = Account.objects.filter(email = email)
        if not user_obj.exists():
            messages.error(request, "Account doesnot exists")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        
        user_obj = auth.authenticate(email=email, password=password)

        if user_obj and user_obj.is_superadmin:
            auth.login(request, user_obj)
            return redirect('admin_dashboard')
        
        messages.error(request, "Invalid Credentials")
        return render(request, 'admin/admin_login.html')

    return render(request, 'admin/admin_login.html')


@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def admin_dashboard(request):
    products = Product.objects.all()
    total_stock  = 0
    for product in products:
        stock = product.total_stock()
        total_stock += stock
    users = Account.objects.filter(is_superadmin=False)
    total_user =users.count()
    pending_delivery = Order.objects.filter(Q(status='Ordered') | Q(status='shipped') | Q(status='Processing')).count()
    request_refunds = Order.objects.filter(refund_requested=True, refund_granted=False)
    request_refunds_count = request_refunds.count()
    orders = Order.objects.filter(status='Delivered')
    total_order = orders.count()
    revenue = orders.aggregate(total_revenue=Sum('order_total'))['total_revenue']
    orders = Order.objects.filter(Q(status='Ordered') | Q(status='shipped') | Q(status='Processing')).order_by('-created_at')[:4]
    if revenue is not None:
        revenue = round(revenue, 2)
    else:
        revenue = 0
    labels = []
    data = []
    sales = (
        Order.objects.filter(status='Delivered')
        .annotate(day=ExtractDay('created_at'))
        .annotate(month=ExtractMonth('created_at'))
        .values('day', 'month')
        .annotate(total_orders=Count('order_total'))
        .annotate(total_sales=Sum('order_total'))
        .order_by('day')
    )
    for i in sales:
        labels.append(f'{i["month"]} {i["day"]}')
        data.append(i["total_orders"])
    context = {
        'request_refunds':request_refunds,
        'request_refunds_count':request_refunds_count,
        'pending_delivery':pending_delivery,
        'total_stock':total_stock,
        'total_user':total_user,
        'orders':orders,
        'users':users,
        'total_order':total_order,
        'revenue':revenue,
        'labels_json': json.dumps(labels),
        'data_json': json.dumps(data),
    }
    return render(request, 'admin/admin_dashboard.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
def sales_report(request):
    if request.method == 'POST':
        from_date = request.POST.get('fromDate')
        to_date = request.POST.get('toDate')
        type = request.POST.get('type')
        order_products= OrderProduct.objects.filter(order__status='Delivered', order__created_at__gte=from_date, order__created_at__lte=to_date).order_by('-created_at')
        order = Order.objects.filter(status="Delivered", created_at__gte=from_date, created_at__lte=to_date)
        total_orders = order.count()
        total_revenue = order.aggregate(total_revenue=Sum('order_total'))['total_revenue']
        total_revenue = round(total_revenue, 2)

        if  request.POST.get('download'):
            print("mai yaha hoon")
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # headers
            headers = ["User", "Product", "Price", "Quantity", "Payment method", "Date"]
            ws.append(headers)

            # data from the model
            for order_product in order_products:
                ws.append([
                    order_product.order.user.email,
                    order_product.product.product_name,
                    order_product.product.price,
                    order_product.quantity,
                    order_product.order.payment_method,
                    order_product.created_at.strftime("%D/%M/%Y") if order_product.created_at else ""
                ])

            dim_holder = DimensionHolder(worksheet=ws)

            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

            ws.column_dimensions = dim_holder

            # Save the workbook to the HttpResponse
            wb.save(response)
            return response
        labels = []
        data = []
        if type == 'weekly':
            sales = (
            Order.objects.filter(status="Delivered", created_at__gte=from_date, created_at__lte=to_date)
            .annotate(week_number=ExtractWeek('created_at'))
            .values('week_number')
            .annotate(total_orders=Count('order_total'))
            .annotate(total_sales=Sum('order_total'))
            .order_by('week_number')
            )
            for i in sales:
                labels.append(f'Week {i["week_number"]}')
                data.append(i["total_orders"])
        elif type == 'monthly':
            sales = (
            Order.objects.filter(status="Delivered", created_at__gte=from_date, created_at__lte=to_date)
            .annotate(month_number=ExtractMonth('created_at'))
            .values('month_number')
            .annotate(total_orders=Count('order_total'))
            .annotate(total_sales=Sum('order_total'))
            .order_by('month_number')
            )
            for i in sales:
                labels.append(f'Month {i["month_number"]}')
                data.append(i["total_orders"])
        else:
            sales = (
            Order.objects.filter(status="Delivered", created_at__gte=from_date, created_at__lte=to_date)
            .annotate(year=ExtractYear('created_at'))
            .values('year')
            .annotate(total_orders=Count('order_total'))
            .annotate(total_sales=Sum('order_total'))
            .order_by('year')
            )
            for i in sales:
                labels.append(f'Year {i["year"]}')
                data.append(i["total_orders"])
        context = {
            'total_orders':total_orders,
            'total_revenue':total_revenue,
            'from_date':from_date,
            'to_date':to_date,
            'order_products':order_products,
            'sales':sales,
            'type':type,
            'labels_json': json.dumps(labels),
            'data_json': json.dumps(data),
        }
        return render(request, 'admin/sales_report.html', context)
    return render(request, 'admin/sales_report.html')


@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def admin_logout(request):
    auth.logout(request)
    messages.info(request, "You are Logged out")
    return redirect('admin_login')

# @user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
# @login_required(login_url='/customadmin/admin_login')
# @never_cache
def user_management(request):    
    users = Account.objects.filter(is_superadmin=False)
    total_user =users.count()
    context = {
        'total_user':total_user,
        'users':users
    }
    return render(request, 'admin/user_management.html', context)
@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def profile(request, pk):
    try:
        profile = Account.objects.get(pk=pk)
    except Exception as e :
        raise e
    context = {
        'profile':profile
    }
    return render(request,'admin/profile.html', context)

# @user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
# @login_required(login_url='/customadmin/admin_login')
# @never_cache
def block_user(request, pk):
    user = Account.objects.get(pk=pk)
    user.is_active=False
    user.is_blocked=True
    user.save()
    return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

# @user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
# @login_required(login_url='/customadmin/admin_login')
# @never_cache
def unblock_user(request, pk):
    user = Account.objects.get(pk=pk)
    user.is_active=True
    user.is_blocked=False
    user.save()
    return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def product(request):
    products = Product.objects.all().order_by('-id')
    context = {
        'products':products,
    }
    return render(request,'admin/product.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def color(request):
    colors = Color.objects.all().order_by('-id')
    context = {
        'colors':colors,
    }
    return render(request,'admin/color.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def size(request):
    sizes = Size.objects.all().order_by('-id')
    context = {
        'sizes':sizes,
    }
    return render(request,'admin/size.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def variation(request):
    variations = Variation.objects.all().order_by('-id')
    context = {
        'variations':variations,
    }
    return render(request,'admin/variation.html', context)

def prod_gallery(request):
    products = Product.objects.all().order_by('-id')
    product_gallery = ProductGallery.objects.all().order_by('-product')
    context = {
        'products':products,
        'product_gallery':product_gallery
    }
    return render(request, 'admin/product_gallery.html', context)

def add_product_gallery(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)
        if form.is_valid():
            image_instance = form.save(commit=False)
            image_instance.product = product
            image_instance.save()
            return JsonResponse({'message': 'works', 'img_id': pk})
        else:
            print(form.errors)
            return JsonResponse({'message': 'error', 'errors': form.errors}, status=400)
    else:
        form = ImageForm()

    return render(request, 'admin/add_product_gallery.html', {'form': form, 'product': product})

def delete_image(request, pk):
    image = ProductGallery.objects.get(pk=pk)
    image.delete()
    messages.success(request, "Image deleted")
    return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def add_product(request):
    categories = Category.objects.all()
    context = {'categories':categories}
    if request.method=='POST':
        product = Product()
        product_name = request.POST['product_name']
        product.description = request.POST['description']
        product.price = request.POST['price']  
        slug = request.POST['slug']
        category = request.POST['category']
        category_id = Category.objects.get(pk=category)
        product.category = category_id
        
        if len(request.FILES) != 0:
            product.image = request.FILES['image']

        if Product.objects.filter(product_name=product_name):
            messages.error(request, "Product already exists")
            return redirect('/customadmin/add_product')

        if Product.objects.filter(slug=slug):
            messages.error(request, "Product Slug already exists")
            return redirect('/customadmin/add_product')
        else:
            product.product_name = product_name
            product.slug = slug
            product.save()
            return redirect('product')

    return render(request, 'admin/add_product.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def add_color(request):
    if request.method == 'POST':
        color_value = request.POST['color']
        color_code = request.POST['color_code']
        if Color.objects.filter(Q(color=color_value) | Q(color_code=color_code)).exists():
            messages.error(request, 'Color or color code already exists')
            return redirect('add_color')
        else:
            color_instance = Color(color=color_value, color_code=color_code)
            color_instance.save()
            messages.success(request, 'Color added successfully.')
            return redirect('color')
    return render(request, 'admin/add_color.html')

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def add_size(request):
    if request.method == 'POST':
        size = request.POST['size']
        if Size.objects.filter(size__iexact=size):
            messages.error(request, 'Size already exists')
            return redirect('add_size')
        else:
            size_instance = Size(size=size)
            size_instance.save()
            messages.success(request, 'Size added successfully.')
            return redirect('size')
    return render(request, 'admin/add_size.html')

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def add_variation(request):
    form = VariationForm(request.POST, request.FILES)
    variation_form = VariationForm()
    products = Product.objects.all()
    colors = Color.objects.all()
    sizes = Size.objects.all()
    context = {
        'variation_form': variation_form,
        'products': products,
        'colors': colors,
        'sizes': sizes
    }
    if request.method == 'POST':
        if form.is_valid():
            product_id = request.POST['product']
            color_value = request.POST['color']
            size_value = request.POST['size']
            product = Product.objects.get(id=product_id)
            color = Color.objects.get(id=color_value)
            size = Size.objects.get(id=size_value)
            # Check if the variation already exists
            if Variation.objects.filter(product=product, color=color, size=size).exists():
                messages.error(request, "Product Variation already exists")
                return redirect('add_variation')
            else:
                form.save()
                return redirect('variation')    
    return render(request, 'admin/add_variation.html', context)


def edit_variation(request, pk):
    variation = Variation.objects.get(pk=pk)
    variation_form = VariationForm(instance=variation)
    context ={
        'variation':variation,
        'variation_form':variation_form
    }
    if request.method == "POST":
        product = request.POST['product']
        color = request.POST['color']
        size = request.POST['size']
        stock = request.POST['stock']

        if len(request.FILES) != 0:
                variation.image = request.FILES['image']
        
        product_id = Product.objects.get(pk=product)
        color_id = Color.objects.get(pk=color)
        size_id = Size.objects.get(pk=size)
        variation.product = product_id
        variation.color = color_id
        variation.size = size_id
        variation.stock = stock
        variation.save()
        messages.success(request, 'Changes saved successfully')
        return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))


    return render(request, 'admin/edit_variation.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def edit_product(request, pk):
    product = Product.objects.get(pk=pk)
    product_gallery = ProductGallery.objects.filter(product_id=product.id)
    product_form = ProductForm(instance=product)
    context = {
        'product':product,
        'product_form':product_form,
        'product_gallery':product_gallery,
    }
    if request.method=='POST':
        product_name = request.POST['product_name']
        slug = request.POST['slug']
        description = request.POST['description']
        price = request.POST['price']
        stock = request.POST['stock']
        category = request.POST['category']
        
        if len(request.FILES) != 0:
            if len(product.image)>0:
                os.remove(product.image.path)
                product.image = request.FILES['image']

        category_id = Category.objects.get(pk=category)
        product.product_name = product_name
        product.slug = slug
        product.description = description
        product.price = price
        product.stock = stock
        product.category = category_id
        product.save()
        messages.success(request, 'Changes saved successfully')
        return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

    return render(request, 'admin/edit_product.html', context)
    
@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def delete_product(request, pk):
    product = Product.objects.get(pk=pk)
    product.deleted = True
    product.save()
    return redirect('/customadmin/product')

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def undelete_product(request, pk):
    product = Product.objects.get(pk=pk)
    product.deleted = False
    product.save()
    return redirect('/customadmin/product')


@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def category(request):
    categories = Category.objects.all().order_by('-id')
    context = {
        'categories':categories,
    }
    return render(request,'admin/category.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def add_category(request):
    categories = Category.objects.all()
    context = {'categories':categories}
    if request.method=='POST':
        category = Category()
        category.category_name = request.POST['category_name']
        category.description = request.POST['description']
        slug = request.POST['slug']

        if len(request.FILES) != 0:
            category.cat_image = request.FILES['cat_image']

        if Category.objects.filter(slug=slug):
            messages.error(request, "Category Slug already exists")
            return redirect('/customadmin/add_category')
        else:
            category.slug = slug
            category.save()
            return redirect('category')

    return render(request, 'admin/add_category.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def edit_category(request, pk):
    category = Category.objects.get(pk=pk)
    category_form = CategoryForm(instance=category)
    categories = Category.objects.all()
    
    context = {
        'category_form':category_form
    }
    if request.method=='POST':
        category_name = request.POST['category_name']
        slug = request.POST['slug']        
        description = request.POST['description']            
        category.category_name = category_name
        category.slug = slug
        category.description = description
        if len(request.FILES) != 0:
            if len(category.image)>0:
                os.remove(category.image.path)
                category.image = request.FILES['image']
        category.save()
        messages.success(request, 'Changes saved successfully')
        return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))


    return render(request, 'admin/edit_category.html', context)
    
@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def delete_category(request, pk):
    category = Category.objects.get(pk=pk)
    category.deleted = True
    category.save()
    return redirect('/customadmin/category')

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def undelete_category(request, pk):
    category = Category.objects.get(pk=pk)
    category.deleted = False
    category.save()
    return redirect('/customadmin/category')

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def order_management(request):
    orders = Order.objects.filter(is_ordered=True).order_by('-created_at')
    request_refunds = Order.objects.filter(refund_requested=True, refund_granted=False)
    context = {
        'orders':orders,
        'request_refunds':request_refunds,
    }
    return render(request,'admin/order_management.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def view_order_detail(request, order_id):
    orders = Order.objects.get(order_number=order_id)
    order_product = OrderProduct.objects.filter(order__order_number=order_id)
    subtotal = 0
    for i in order_product:
        subtotal += i.product.price * i.quantity
    context = {
        'orders':orders,
        'order_product':order_product,
        'subtotal':subtotal
    }
    return render(request, 'admin/view_order_detail.html', context) 

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def change_order_status(request, order_id):
    order = Order.objects.get(order_number=order_id)
    user = order.user
    print(user)
    if request.method == "POST":        
        status = request.POST['status']
        order.status = status
        if order.status == 'Cancelled' or order.status == 'refunded':
            if order.payment_method != 'COD':
                wallet = Decimal(str(order.order_total)) 
                user.wallet += wallet
                user.save()
        order.save()
        messages.success(request, "Order status has been updated")   
    else:
        order_form = OrderForm(instance=request.user)
    order_form = OrderForm(instance=order)
    context = {
            'orders':order,
            'order_form':order_form,
        }
       
    return render(request, 'admin/change_order_status.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def admin_cancel_order(request, order_id):
    print('I was called')
    order = Order.objects.get(order_number=order_id)
    if order.status:
        print('I was here')
        order.status = 'Cancelled'
        order.save()
        messages.success(request, "Order Cancelled Successfully.")
    return redirect('order_management')


@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def view_return_order(request, order_id):
    orders = Order.objects.get(order_number=order_id)
    order_product = OrderProduct.objects.filter(order__order_number=order_id)
    refund = Refund.objects.get(order=orders)
    subtotal = 0
    for i in order_product:
        subtotal += i.product.price * i.quantity
    context = {
        'orders':orders,
        'order_product':order_product,
        'subtotal':subtotal,
        'refund':refund,
    }
    return render(request, 'admin/view_return_order.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def admin_grant_return_request(request, order_id):
    order = Order.objects.get(order_number=order_id)
    user = Account.objects.get(id=order.user.id)
    if order.refund_requested == True and order.refund_granted == False:
        order.refund_granted = True
        order.status = 'refunded'
        wallet = Decimal(str(order.order_total)) 
        user.wallet += wallet
        user.save()
        order.save()
    return redirect('order_management')

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def coupon(request):
    coupons = Coupon.objects.all()
    context = {
        'coupons':coupons,
    }
    return render(request,'admin/coupon.html', context)

@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def add_coupon(request):
    if request.method == "POST":
        coupon = Coupon()
        coupon_code = request.POST['coupon_code']
        if Coupon.objects.filter(coupon_code=coupon_code):
            messages.error(request, "Coupon already exists")
            return redirect('add_coupon')
        else:
            coupon.coupon_code = coupon_code
            coupon.minimum_amount = request.POST['minimum_amount']
            coupon.discount = request.POST['discount']        
            coupon.save()
            return redirect('coupon')
    else:
        return render(request, 'admin/add_coupon.html')
    
@user_passes_test(user_is_admin, login_url='/customadmin/admin_login')
@login_required(login_url='/customadmin/admin_login')
@never_cache
def edit_coupon(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    if request.method == 'POST':
        coupon.coupon_code = request.POST.get('coupon_code')
        coupon.discount = request.POST.get('discount')
        coupon.minimum_amount = request.POST.get('minimum_amount', 0)
        coupon.save()
        messages.success(request, 'Coupon updated successfully!')
        return redirect('coupon')
    return render(request, 'admin/edit_coupon.html', {'coupon': coupon})

def undelete_coupon(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.is_expired = False
    coupon.save()
    return redirect('coupon')

def delete_coupon(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.is_expired = True
    coupon.save()
    return redirect('coupon')


    





